import streamlit as st
import pandas as pd
import openpyxl
import re
from datetime import date, timedelta, datetime
from io import BytesIO
import numpy as np
import matplotlib.pyplot as plt
# --- Importations spécifiques à Gemini ---
import google.genai as genai
from google.genai.errors import APIError
from google.genai import types 
import json
import hashlib


# Vérification optionnelle de la dépendance 'tabulate' (utile pour to_markdown dans le contexte IA)
try:
    import tabulate
except ImportError:
    pass


# ---------------------------
# Fonction pour fixer le chat input en bas et styliser les bulles de chat
# ---------------------------
def apply_custom_styles():
    """
    Injecte du CSS pour fixer le champ de requête IA en bas de l'écran (dans les limites du conteneur principal)
    ET pour styliser les messages en bulles de chat distinctes.
    """
    st.markdown(
        """
        <style>
        /* 1. FIXER ET REDIMENSIONNER LE CHAT INPUT */
        /* Cible le conteneur principal du chat_input */
        .stChatInput {
            position: fixed;
            bottom: 0px;
            /* Utilise le conteneur principal de Streamlit pour limiter la largeur */
            /* La largeur réelle sera gérée par le conteneur dans le bloc-container */
            left: 0;
            right: 0; 
            z-index: 9999;
            padding: 10px 0;
            background-color: white; 
            box-shadow: 0 -2px 5px rgba(0, 0, 0, 0.1); /* Ombre légère pour le démarquer */
        }
        
        /* Assure que le chat input est aligné avec le contenu du bloc-container */
        /* Cible le bloc container dans la section main (où le contenu est affiché) */
        section.main .block-container {
             /* Ajout d'un padding pour que le contenu ne soit pas masqué par la barre fixe */
             padding-bottom: 90px; 
        }

        /* Ajuste la largeur du chat input pour qu'il ne déborde pas du contenu */
        /* Cible la div Streamlit qui contient le chat_input pour centrer et limiter */
        div.stChatInput > div > div {
             max-width: 900px; /* Limiter la largeur max comme le contenu normal */
             margin: 0 auto; /* Centrer dans la zone disponible */
             padding: 0 1rem; /* Marge intérieure */
        }


        /* 2. STYLISATION DES BULLES DE CHAT */

        /* Styles de base pour toutes les bulles de chat */
        .stChatMessage {
            margin-bottom: 10px;
        }

        /* Message de l'utilisateur (bleu à droite) */
        /* Cible le conteneur du message de l'utilisateur (role="user") */
        .stChatMessage[data-testid="stChatMessage"]:nth-child(odd) > div:first-child > div:nth-child(2) {
            background-color: #E0F2F1; /* Couleur plus claire pour l'utilisateur */
            border-radius: 15px 15px 0 15px; /* Angle vif en bas à gauche */
            padding: 10px 15px;
            margin-left: auto; /* Pousser la bulle vers la droite */
            max-width: 80%;
            text-align: left;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        .stChatMessage[data-testid="stChatMessage"]:nth-child(odd) { /* Pour aligner l'avatar/nom à droite */
             flex-direction: row-reverse; 
        }
        .stChatMessage[data-testid="stChatMessage"]:nth-child(odd) .st-emotion-cache-1r5j5s6 { /* Cible l'avatar user */
            margin-left: 10px;
            margin-right: 0;
        }


        /* Message de l'assistant (vert à gauche) */
        /* Cible le conteneur du message de l'assistant (role="assistant") */
        .stChatMessage[data-testid="stChatMessage"]:nth-child(even) > div:first-child > div:nth-child(2) {
            background-color: #F0F0F0; /* Couleur neutre/grise */
            border-radius: 15px 15px 15px 0; /* Angle vif en bas à droite */
            padding: 10px 15px;
            margin-right: auto; /* Pousser la bulle vers la gauche */
            max-width: 80%;
            text-align: left;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }
        .stChatMessage[data-testid="stChatMessage"]:nth-child(even) .st-emotion-cache-1r5j5s6 { /* Cible l'avatar assistant */
            margin-right: 10px;
            margin-left: 0;
        }

        
        
        </style>
        """,
        unsafe_allow_html=True
    )

# ---------------------------
# Configuration de la page Streamlit
# ---------------------------
st.set_page_config(page_title="Synthèse Totaux SAGE | ERP", layout="wide")
st.title("📊 Synthèse SAGE | ERP")

# Appeler la fonction de fix et de styles dès le début
apply_custom_styles()

# --- Configuration de l'API Gemini ---
GEMINI_API_KEY = None
MODEL_NAME = None
CLIENT_KEY = "gemini_client" 

try:
    if "GEMINI_API_KEY" not in st.secrets:
        # Permettre à l'app de continuer sans IA
        pass 
    else:
        GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
        MODEL_NAME = "gemini-2.5-flash"
        
        if CLIENT_KEY not in st.session_state or st.session_state[CLIENT_KEY] is None:
            st.session_state[CLIENT_KEY] = genai.Client(api_key=GEMINI_API_KEY)

except Exception as e:
    # Gérer l'échec de l'initialisation du client sans bloquer le reste de l'application
    MODEL_NAME = None 

# ---------------------------
# Fonctions utilitaires 
# ---------------------------
# (Fonctions parse_excel_number_like, fmt_number, fmt_ecart_gap, _to_num_display, extract_last_total_row, is_date_format, extract_all_dates_from_file, compute_totals, create_pie_chart, create_bar_chart, render_html_table, write_styled_table, calculate_detailed_gaps, display_reconciliation_table, generate_full_context, generate_analysis_id, get_total_summary restent inchangées)

# [... toutes les fonctions utilitaires restent ici : parse_excel_number_like, fmt_number, fmt_ecart_gap, _to_num_display, extract_last_total_row, is_date_format, extract_all_dates_from_file, compute_totals, create_pie_chart, create_bar_chart, render_html_table, write_styled_table, calculate_detailed_gaps, display_reconciliation_table, generate_full_context, generate_analysis_id, get_total_summary ...]

# Pour la concision du code affiché, nous omettons les corps de fonctions non modifiées, mais elles doivent être présentes dans le fichier de l'utilisateur.
# Ici, nous ne modifions que les fonctions de persistence et d'affichage de l'historique.

def parse_excel_number_like(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except:
            return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace("\u202f", "").replace("\xa0", "").replace("\u2009", "")
    s = s.replace("−", "-")
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        elif "." in s and s.count(".") > 1:
            s = s.replace(".", "")
    s_clean = re.sub(r"[^\d\.-]", "", s)
    if s_clean in ("", "-", ".", "-."):
        return None
    try:
        return float(s_clean)
    except:
        return None

def fmt_number(v):
    try:
        if v is None:
            return ""
        fv = float(v)
    except:
        return ""
    if abs(fv) < 0.01:
        return "0,00"
    try:
        return f"{fv:,.2f}".replace(",", "_TEMP_").replace(".", ",").replace("_TEMP_", " ")
    except:
        return ""

def fmt_ecart_gap(v):
    try:
        if v is None:
            return ""
        fv = float(v)
    except:
        return ""
    if abs(fv) < 0.01:
        return ""
    try:
        if isinstance(v, str) and "%" in v:
            return v
        return f"{fv:,.2f}".replace(",", "_TEMP_").replace(".", ",").replace("_TEMP_", " ")
    except:
        return ""

def _to_num_display(s):
    if not s:
        return 0.0
    ss = str(s)
    ss = ss.replace("\u202f", "").replace("\xa0", "").replace(" ", "").replace(",", ".")
    ss = ss.strip()
    try:
        return float(ss)
    except:
        return 0.0

def extract_last_total_row(uploaded_file):
    try:
        file_extension = uploaded_file.name.split('.')[-1].lower()
        # Logique CSV/XLSX
        if file_extension == 'xlsx':
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
        elif file_extension == 'csv':
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding='latin-1', on_bad_lines='skip', low_memory=False)
            last_total_candidates = []
            for idx, row in df.iterrows():
                if row.astype(str).str.contains("total", case=False).any():
                    last_total_candidates.append(idx)
            if not last_total_candidates:
                return []
            total_row_idx = last_total_candidates[-1]
            numeric_values = []
            for val in df.iloc[total_row_idx]:
                parsed = parse_excel_number_like(val)
                if parsed is not None:
                    numeric_values.append(float(parsed))
            uploaded_file.seek(0)
            return numeric_values
        else:
            return []

        # Logique pour XLSX (OpenPyXL)
        last_total_candidates = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value_str = cell.value.strip().lower()
                    if any(x in value_str for x in ["total", "totaux", "total général"]):
                        is_bold = cell.font and cell.font.bold
                        last_total_candidates.append((cell.row, bool(is_bold)))
        if not last_total_candidates:
            return []
        last_bold = [r for r in reversed(last_total_candidates) if r[1]]
        total_row_idx = last_bold[0][0] if last_bold else last_total_candidates[-1][0]
        numeric_values = []
        for cell in ws[total_row_idx]:
            raw = cell.value
            parsed = parse_excel_number_like(raw)
            if parsed is not None:
                numeric_values.append(float(parsed))
        return numeric_values

    except Exception as e:
        return []

def is_date_format(s, fmt):
    try:
        datetime.strptime(s, fmt)
        return True
    except (ValueError, TypeError):
        return False

def extract_all_dates_from_file(uploaded_file):
    if not uploaded_file:
        return None, None, None
    date_debut, date_fin, extraction_date = None, None, None
    try:
        uploaded_file.seek(0)
        # Logique d'extraction de date...
        if uploaded_file.name.endswith('.csv'):
            content = uploaded_file.getvalue().decode('latin-1', errors='ignore')
            for line in content.split('\n')[:20]:
                matches = re.findall(r'\b(\d{2}/\d{2}/\d{4})\b', line)
                if len(matches) == 2:
                    try:
                        date_debut = datetime.strptime(matches[0], "%d/%m/%Y").date()
                        date_fin = datetime.strptime(matches[1], "%d/%m/%Y").date()
                    except ValueError:
                        pass
                if not extraction_date:
                    match_extract = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', line)
                    if match_extract and not any(k in line.lower() for k in ["solde", "ouverture", "mouvement", "cumulé"]):
                        extraction_date = datetime.strptime(match_extract.group(1), "%d/%m/%Y").date()

        elif uploaded_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        lower_val = cell.value.lower()
                        if "période du" in lower_val and date_debut is None:
                            start_date_cell_val = ws.cell(row=cell.row, column=cell.column + 1).value
                            if isinstance(start_date_cell_val, datetime):
                                date_debut = start_date_cell_val.date()
                        if "au" == lower_val and date_fin is None:
                            end_date_cell_val = ws.cell(row=cell.row, column=cell.column + 1).value
                            if isinstance(end_date_cell_val, datetime):
                                date_fin = end_date_cell_val.date()
                        if "date de tirage" in lower_val and extraction_date is None:
                            extraction_date_cell_val = ws.cell(row=cell.row, column=cell.column + 1).value
                            if isinstance(extraction_date_cell_val, datetime):
                                extraction_date = extraction_date_cell_val.date()

        uploaded_file.seek(0)

    except Exception as e:
        pass
    return date_debut, date_fin, extraction_date

def compute_totals(entity_name, erp_file=None, sage_file=None):
    sage_nums_raw = [0.0] * 6
    erp_nums_raw = [0.0] * 6
    # Logique de calcul des totaux...
    if sage_file:
        sage_totaux = extract_last_total_row(sage_file)
        if len(sage_totaux) >= 6:
            sage_nums_raw = [float(v) for v in sage_totaux[:6]]
        elif len(sage_totaux) >= 4:
            d_init, c_init, mvt_d, mvt_c = sage_totaux[:4]
            solde_net = (d_init + mvt_d) - (c_init + mvt_c)
            d_fin = max(0, solde_net)
            c_fin = max(0, -solde_net)
            sage_nums_raw = [d_init, c_init, mvt_d, mvt_c, d_fin, c_fin]

    if erp_file:
        erp_totaux_all = extract_last_total_row(erp_file)
        if len(erp_totaux_all) >= 6:
            erp_nums_raw = [float(v) for v in erp_totaux_all[:6]]
        elif len(erp_totaux_all) >= 4:
            d_init, c_init, mvt_d, mvt_c = erp_totaux_all[:4]
            solde_net = (d_init + mvt_d) - (c_init + mvt_c)
            d_fin = max(0, solde_net)
            c_fin = max(0, -solde_net)
            erp_nums_raw = [d_init, c_init, mvt_d, mvt_c, d_fin, c_fin]

    def normalize_balance(d_val, c_val):
        if d_val > 0 and c_val > 0:
            if d_val > c_val:
                return d_val - c_val, 0.0
            else:
                return 0.0, c_val - d_val
        return d_val, c_val

    sage_nums_raw[0], sage_nums_raw[1] = normalize_balance(sage_nums_raw[0], sage_nums_raw[1])
    erp_nums_raw[0], erp_nums_raw[1] = normalize_balance(erp_nums_raw[0], erp_nums_raw[1])
    sage_nums_raw[4], sage_nums_raw[5] = normalize_balance(sage_nums_raw[4], sage_nums_raw[5])
    erp_nums_raw[4], erp_nums_raw[5] = normalize_balance(erp_nums_raw[4], erp_nums_raw[5])

    sage_values_fmt = [fmt_number(v) for v in sage_nums_raw]
    erp_values_fmt = [fmt_number(v) for v in erp_nums_raw]

    return sage_nums_raw, sage_values_fmt, erp_nums_raw, erp_values_fmt

def create_pie_chart(title, values, labels):
    fig, ax = plt.subplots(figsize=(8, 8))
    non_zero_values = [v for v in values if v > 0]
    non_zero_labels = [l for v, l in zip(values, labels) if v > 0]
    if not non_zero_values:
        return
    def absolute_value(val):
        return f"{val:.1f}%"
    wedges, texts, autotexts = ax.pie(non_zero_values, autopct=absolute_value, startangle=90,
                                       wedgeprops={'linewidth': 1.0, 'edgecolor': 'white'})
    ax.legend(wedges, non_zero_labels, title="Légende", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
    ax.axis('equal')
    st.pyplot(fig)

def create_bar_chart(title, labels, sage_values, erp_values, diff_values, gap_values):
    """Crée et affiche un graphique à barres groupées pour SAGE, ERP, et l'écart."""
    sage_nets = [
        abs(sage_values[0] - sage_values[1]),
        abs(sage_values[2] - sage_values[3]),
        abs(sage_values[4] - sage_values[5])
    ]
    erp_nets = [
        abs(erp_values[0] - erp_values[1]),
        abs(erp_values[2] - erp_values[3]),
        abs(erp_values[4] - erp_values[5])
    ]
    ecarts_nets = [s - e for s, e in zip(sage_nets, erp_nets)]
    fig, ax = plt.subplots(figsize=(12, 7))
    x_labels = ["Solde d'ouverture", "Mouvements", "Solde Final"]
    x = np.arange(len(x_labels))
    width = 0.30
    offset = 0.20
    width_diff = 0.20
    rects1 = ax.bar(x - offset, sage_nets, width, label='SAGE', color='#3498DB', zorder=2)
    rects2 = ax.bar(x + offset, erp_nets, width, label='ERP', color='#2ECC71', zorder=2)
    rects3 = ax.bar(x, ecarts_nets, width_diff, label='Écart', color='#E74C3C', alpha=0.8, zorder=3)
    all_values = sage_nets + erp_nets + ecarts_nets
    min_val = min(all_values) if all_values else 0
    max_val = max(all_values) if all_values else 0
    abs_max = max(abs(min_val), abs(max_val)) * 1.1 if max(abs(min_val), abs(max_val)) > 0 else 1000
    ax.set_ylim(-abs_max, abs_max)
    def format_y(y, pos):
        return f"{y:,.0f}".replace(",", " ").replace(".", ",")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(format_y))
    ax.set_ylabel('Valeurs NETTES (FCFA)', color='black')
    ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(x_labels, rotation=15, ha="right")
    ax.axhline(0, color='black', linewidth=1.2, linestyle='-', zorder=1)
    lines, labels_ = ax.get_legend_handles_labels()
    ax.legend(lines, labels_, loc='upper left')
    fig.tight_layout()
    st.pyplot(fig)


def render_html_table(entity_name, erp_file, sage_file, extraction_date, date_debut, date_fin, sage_nums=None, erp_nums=None, ecarts_raw=None):
    
    # Si les données ne sont pas fournies (mode historique), elles sont recalculées ou déduites des fichiers
    if sage_nums is None or erp_nums is None:
        sage_nums, sage_values_fmt, erp_nums, erp_values_fmt = compute_totals(entity_name, erp_file, sage_file)
    else:
        # Utilise les données passées (cas historique)
        sage_values_fmt = [fmt_number(v) for v in sage_nums]
        erp_values_fmt = [fmt_number(v) for v in erp_nums]
        
    if ecarts_raw is None:
        ecarts_raw = [(s - e) for s, e in zip(sage_nums, erp_nums)]
        
    ecarts_fmt = [fmt_ecart_gap(x) for x in ecarts_raw[:4]] + ["", ""]
    gaps_raw = [f"{(e / s) * 100:.2f}%" if s != 0 and abs(s) >= 0.01 else fmt_ecart_gap(0) for e, s in zip(ecarts_raw[:4], sage_nums[:4])]
    gaps = [g for g in gaps_raw] + ["", ""]
    
    headers_display = [
        f"Données tirée le {extraction_date.strftime('%d/%m/%Y')}",
        "Environnement",
        f"Solde d'ouverture<br>au {date_debut.strftime('%d/%m/%Y')}",
        f"Mouvement du {date_debut.strftime('%d/%m/%Y')}<br>au {date_fin.strftime('%d/%m/%Y')}",
        f"Solde au<br>{date_fin.strftime('%d/%m/%Y')}"
    ]
    sub_headers = ["Débit", "Crédit"] * 3
    custom_css = f"""
    <style>
      .totaux-table {{ border-collapse: collapse; width: 100%; table-layout: fixed; font-family: Arial, sans-serif; }}
      .totaux-table th.main-header {{ background-color: #e9ecef; padding: 10px; text-align: center; font-weight: bold; border: 1px solid #adb5bd; height: 45px; vertical-align: middle; }}
      .totaux-table th.sub-header {{ background-color: #f8f9fa; padding: 5px; text-align: center; font-weight: bold; border: 1px solid #adb5bd; font-size: 0.85em; }}
      .totaux-table td {{ padding: 10px; text-align: center; vertical-align: middle; word-wrap: break-word; border: 1px solid #ddd; }}
      .totaux-table .bloc-titre-cell {{ font-weight: bold; background-color: #e9ecef; }}
      .totaux-table .row-ecart td {{ font-weight: bold; }}
      .totaux-table .row-ecart td:nth-child(2) {{ color: #e67e22; }}
      .totaux-table .row-gap td {{ background-color: #f8d7da; color: #721c24; font-weight: bold; }}
      .totaux-table .row-gap td:first-child {{ background-color: transparent; }}
    </style>
    """
    html_table = f"""
    <table class="totaux-table">
      <tr>
        <th class="main-header" rowspan="2">{headers_display[0]}</th>
        <th class="main-header" rowspan="2">{headers_display[1]}</th>
        <th class="main-header" colspan="2">{headers_display[2]}</th>
        <th class="main-header" colspan="2">{headers_display[3]}</th>
        <th class="main-header" colspan="2">{headers_display[4]}</th>
      </tr>
      <tr>
        <th class="sub-header">{sub_headers[0]}</th><th class="sub-header">{sub_headers[1]}</th>
        <th class="sub-header">{sub_headers[2]}</th><th class="sub-header">{sub_headers[3]}</th>
        <th class="sub-header">{sub_headers[4]}</th><th class="sub-header">{sub_headers[5]}</th>
      </tr>
      <tr>
        <td class="bloc-titre-cell" rowspan="2">Solde-Balance Auxiliaire<br>{entity_name}</td>
        <td>SAGE</td>
        <td>{sage_values_fmt[0]}</td><td>{sage_values_fmt[1]}</td>
        <td>{sage_values_fmt[2]}</td><td>{sage_values_fmt[3]}</td>
        <td>{sage_values_fmt[4]}</td><td>{sage_values_fmt[5]}</td>
      </tr>
      <tr>
        <td>ERP</td>
        <td>{erp_values_fmt[0]}</td><td>{erp_values_fmt[1]}</td>
        <td>{erp_values_fmt[2]}</td><td>{erp_values_fmt[3]}</td>
        <td>{erp_values_fmt[4]}</td><td>{erp_values_fmt[5]}</td>
      </tr>
      <tr class="row-ecart">
        <td></td>
        <td>Ecart</td>
        <td>{ecarts_fmt[0]}</td><td>{ecarts_fmt[1]}</td>
        <td>{ecarts_fmt[2]}</td><td>{ecarts_fmt[3]}</td>
        <td>{ecarts_fmt[4]}</td><td>{ecarts_fmt[5]}</td>
      </tr>
      <tr class="row-gap">
        <td></td>
        <td>Gap</td>
        <td>{gaps[0]}</td><td>{gaps[1]}</td>
        <td>{gaps[2]}</td><td>{gaps[3]}</td>
        <td>{gaps[4]}</td><td>{gaps[5]}</td>
      </tr>
    </table>
    """
    st.markdown(custom_css + html_table, unsafe_allow_html=True)
    return sage_nums, erp_nums, ecarts_raw, gaps

def write_styled_table(ws, start_row, entity_name, erp_file, sage_file, extraction_date, date_debut, date_fin, section_title=None):
    sage_nums, sage_values_fmt, erp_nums, erp_values_fmt = compute_totals(entity_name, erp_file, sage_file)
    ecarts_raw = [(s - e) for s, e in zip(sage_nums, erp_nums)]
    # Logique de création de tableau Excel... (inchangée)
    ecarts_fmt = [fmt_ecart_gap(x) for x in ecarts_raw[:4]] + ["", ""]
    gaps_raw_values = [f"{(e / s) * 100:.2f}%" if s != 0 and abs(s) >= 0.01 else fmt_ecart_gap(0) for e, s in zip(ecarts_raw[:4], sage_nums[:4])]
    gaps = [g for g in gaps_raw_values] + ["", ""]
    headers_main = [
        f"Données tirée le {extraction_date.strftime('%d/%m/%Y')}",
        "Environnement",
        f"Solde d'ouverture\nau {date_debut.strftime('%d/%m/%Y')}",
        f"Mouvement du {date_debut.strftime('%d/%m/%Y')}\nau {date_fin.strftime('%d/%m/%Y')}",
        f"Solde au\n{date_fin.strftime('%d/%m/%Y')}"
    ]
    sub_headers = ["Débit", "Crédit"] * 3
    rows = [
        [f"Solde-Balance Auxiliaire\n{entity_name}", "SAGE"] + list(sage_values_fmt),
        ["", "ERP"] + list(erp_values_fmt),
        ["", "Ecart"] + list(ecarts_fmt),
        ["", "Gap"] + list(gaps),
    ]
    thin = openpyxl.styles.Side(style="thin", color="BBBBBB")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill_main = openpyxl.styles.PatternFill("solid", fgColor="E9ECEF")
    header_fill_sub = openpyxl.styles.PatternFill("solid", fgColor="F8F9FA")
    gap_fill = openpyxl.styles.PatternFill("solid", fgColor="F8D7DA")
    gap_label_fill = openpyxl.styles.PatternFill("solid", fgColor="F5C6CB")
    gap_font = openpyxl.styles.Font(bold=True, color="721C24")
    ecart_font = openpyxl.styles.Font(bold=True, color="E67E22")
    center_wrap = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    if section_title:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        t = ws.cell(row=start_row, column=1, value=section_title)
        t.font = openpyxl.styles.Font(bold=True, size=14)
        t.alignment = center
        start_row_main_header = start_row + 2
    else:
        start_row_main_header = start_row
    start_row_sub_header = start_row_main_header + 1
    ws.merge_cells(start_row=start_row_main_header, start_column=1, end_row=start_row_sub_header, end_column=1)
    ws.merge_cells(start_row=start_row_main_header, start_column=2, end_row=start_row_sub_header, end_column=2)
    for col_start in [3, 5, 7]:
        ws.merge_cells(start_row=start_row_main_header, start_column=col_start, end_row=start_row_main_header, end_column=col_start + 1)
    main_cols = [1, 2, 3, 5, 7]
    for col_idx, value in zip(main_cols, headers_main):
        c = ws.cell(row=start_row_main_header, column=col_idx, value=value)
        c.font = openpyxl.styles.Font(bold=True)
        c.fill = header_fill_main
        c.alignment = center_wrap
        c.border = border
    for col, value in enumerate(sub_headers, start=3):
        c = ws.cell(row=start_row_sub_header, column=col, value=value)
        c.font = openpyxl.styles.Font(bold=True, size=10)
        c.fill = header_fill_sub
        c.alignment = center
        c.border = border
    start_data_row = start_row_sub_header + 1
    ws.merge_cells(start_row=start_data_row, start_column=1, end_row=start_data_row + 1, end_column=1)
    for i, row_vals in enumerate(rows, start=0):
        for col, value in enumerate(row_vals, start=1):
            if i == 1 and col == 1:
                continue
            c = ws.cell(row=start_data_row + i, column=col, value=value)
            if i < 2 and col >= 3 and value:
                try:
                    num_value = _to_num_display(value)
                    c.value = num_value
                    c.number_format = '#,##0.00'
                except:
                    pass
            c.alignment = center
            c.border = border
            if i == 0 and col == 1:
                c.font = openpyxl.styles.Font(bold=True)
                c.alignment = center_wrap
            if i == 2 and col == 2:
                c.font = ecart_font
            elif i == 2 and col >= 3 and value:
                 try:
                    num_value = _to_num_display(value)
                    c.value = num_value
                    c.number_format = '#,##0.00'
                 except:
                    pass
            if i == 3:
                c.fill = gap_fill
                c.font = gap_font
                if col == 2:
                    c.fill = gap_label_fill
                    c.font = gap_font
                if col >= 3 and col <= 6 and value:
                    c.number_format = '0.00%'
    widths = [28, 14] + [15] * 6
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.row_dimensions[start_row_main_header].height = 20
    ws.row_dimensions[start_row_sub_header].height = 20
    ws.row_dimensions[start_data_row].height = 25
    ws.row_dimensions[start_data_row + 1].height = 25
    ws.row_dimensions[start_data_row + 2].height = 25
    ws.row_dimensions[start_data_row + 3].height = 25
    return start_data_row + 4


def calculate_detailed_gaps(df_correspondances, entity_type="Fournisseur"):
    if df_correspondances.empty:
        return pd.DataFrame()
    account_col = df_correspondances.columns[0]
    name_col = df_correspondances.columns[1]
    col_prefix = "N° Cpte" if entity_type == "Client" else "N° Cpte Fournisseur"
    final_columns = [
        f"{col_prefix} SAGE/ERP",
        "Désignation",
    ]
    num_cols = ["1", "2", "3"]
    dc_cols = ["Débit", "Crédit"]
    df_result = pd.DataFrame()
    df_result[final_columns[0]] = df_correspondances[account_col].astype(str)
    df_result[final_columns[1]] = df_correspondances[name_col].astype(str)
    for num in num_cols:
        for dc in dc_cols:
            col_name_base = f"{dc} {num}"
            sage_col = f"{col_name_base}.SAGE"
            erp_col = f"{col_name_base}.ERP"
            if sage_col in df_correspondances.columns and erp_col in df_correspondances.columns:
                sage_data = pd.to_numeric(df_correspondances[sage_col], errors='coerce').fillna(0)
                erp_data = pd.to_numeric(df_correspondances[erp_col], errors='coerce').fillna(0)
                ecart_raw = sage_data - erp_data
                new_sage_col_name = f"{col_name_base} SAGE"
                new_erp_col_name = f"{col_name_base} ERP"
                new_ecart_col_name = f"Écart {col_name_base}"
                df_result[new_sage_col_name] = sage_data.apply(fmt_number)
                df_result[new_erp_col_name] = erp_data.apply(fmt_number)
                df_result[new_ecart_col_name] = ecart_raw.apply(fmt_ecart_gap)
    ecart_cols_final = [col for col in df_result.columns if col.startswith('Écart ')]
    if ecart_cols_final:
        # Filtrer uniquement les lignes où il y a un écart non nul (formaté non vide)
        df_result = df_result[df_result[ecart_cols_final].apply(lambda row: any(row.astype(str).str.len() > 0), axis=1)]
    ordered_cols = df_result.columns[:2].tolist()
    for num in num_cols:
        for dc in dc_cols:
            base = f"{dc} {num}"
            ordered_cols.append(f"{base} SAGE")
            ordered_cols.append(f"{base} ERP")
            ordered_cols.append(f"Écart {base}")
    ordered_cols_clean = [col for col in ordered_cols if col in df_result.columns]
    df_result = df_result.reindex(columns=ordered_cols_clean, fill_value="")
    return df_result


def display_reconciliation_table(df_reconciliation, title):
    # st.markdown(f"### {title}") # Le titre est mis dans la section Historique
    if df_reconciliation.empty:
        st.info("Aucune ligne d'écart individuel à afficher après filtrage (ou pas de correspondance trouvée).")
        return
    def highlight_gaps(s):
        is_gap = s.astype(str).str.strip().str.len() > 0
        return [
            'background-color: #f8d7da; font-weight: bold; color: #721c24' if v else ''
            for v in is_gap
        ]
    gap_cols = [col for col in df_reconciliation.columns if col.startswith('Écart ')]
    if gap_cols:
        styled_df = df_reconciliation.style.apply(highlight_gaps, subset=pd.IndexSlice[:, gap_cols])
        st.dataframe(styled_df, height=500, use_container_width=True)
    else:
        st.dataframe(df_reconciliation, height=500, use_container_width=True)

        
def generate_full_context(sage_f_nums, erp_f_nums, ecarts_f, sage_c_nums, erp_c_nums, ecarts_c, df_reco_f=None, df_reco_c=None):
    """Génère le contexte complet de l'analyse pour l'assistant IA en format textuel."""
    
    total_data = {
        "Fournisseurs": {
            "SAGE": sage_f_nums, "ERP": erp_f_nums, "Ecarts_Bruts": ecarts_f,
            "Postes": ["Débit Ouv.", "Crédit Ouv.", "Débit Mvt", "Crédit Mvt", "Débit Fin.", "Crédit Fin."]
        },
        "Clients": {
            "SAGE": sage_c_nums, "ERP": erp_c_nums, "Ecarts_Bruts": ecarts_c,
            "Postes": ["Débit Ouv.", "Crédit Ouv.", "Débit Mvt", "Crédit Mvt", "Débit Fin.", "Crédit Fin."]
        }
    }
    
    # Récupérer les dates de la session pour le contexte IA
    date_debut_ia = st.session_state.get('date_debut_global', date.today())
    date_fin_ia = st.session_state.get('date_fin_global', date.today())
    
    context_text = f"Analyse de réconciliation SAGE vs ERP pour la période du {date_debut_ia.strftime('%d/%m/%Y')} au {date_fin_ia.strftime('%d/%m/%Y')}. Toutes les valeurs sont en FCFA (Francs CFA).\n\n"
    
    # 1. Analyse des Totaux
    for entity, data in total_data.items():
        if all(sum(data[env]) == 0 for env in ["SAGE", "ERP"]):
            context_text += f"Aucune donnée de totaux significative trouvée pour les {entity}.\n\n"
            continue
            
        context_text += f"## Totaux Agrégés pour les {entity}\n"
        
        sage_net_ouv = data['SAGE'][0] - data['SAGE'][1]
        erp_net_ouv = data['ERP'][0] - data['ERP'][1]
        ecart_net_ouv = sage_net_ouv - erp_net_ouv
        
        sage_net_mvt = data['SAGE'][2] - data['SAGE'][3]
        erp_net_mvt = data['ERP'][2] - data['ERP'][3]
        ecart_net_mvt = sage_net_mvt - erp_net_mvt
        
        sage_net_fin = data['SAGE'][4] - data['SAGE'][5]
        erp_net_fin = data['ERP'][4] - data['ERP'][5]
        ecart_net_fin = sage_net_fin - erp_net_fin
        
        # Fonction de formatage pour le contexte (similaire au format FR)
        def fmt_ctx(v):
            # Formate avec un point comme séparateur de milliers et une virgule pour les décimales
            return f"{v:,.2f}".replace(",", "_TEMP_").replace(".", ",").replace("_TEMP_", " ")

        context_text += f"| Poste | SAGE (Net) | ERP (Net) | Écart (SAGE - ERP) |\n"
        context_text += f"| :--- | ---: | ---: | ---: |\n"
        context_text += f"| Solde d'Ouverture Net | {fmt_ctx(sage_net_ouv)} | {fmt_ctx(erp_net_ouv)} | {fmt_ctx(ecart_net_ouv)} |\n"
        context_text += f"| Mouvements Net | {fmt_ctx(sage_net_mvt)} | {fmt_ctx(erp_net_mvt)} | {fmt_ctx(ecart_net_mvt)} |\n"
        context_text += f"| Solde Final Net | {fmt_ctx(sage_net_fin)} | {fmt_ctx(erp_net_fin)} | {fmt_ctx(ecart_net_fin)} |\n\n"
        
        context_text += f"Détail Débit/Crédit:\n"
        for i, poste in enumerate(data['Postes']):
            ecart_raw = data['Ecarts_Bruts'][i]
            gap = f"{(ecart_raw / data['SAGE'][i] * 100):.2f}%" if data['SAGE'][i] != 0 and abs(data['SAGE'][i]) >= 0.01 else "N/A"
            context_text += f"- {poste}: SAGE={fmt_ctx(data['SAGE'][i])}, ERP={fmt_ctx(data['ERP'][i])}, Écart={fmt_ctx(ecart_raw)}, Gap={gap}\n"
        context_text += "\n"

    # 2. Analyse Ligne-à-Ligne
    if df_reco_f is not None and not df_reco_f.empty:
        try:
            # Utilisation de tabulate pour générer un tableau Markdown
            df_f_str = df_reco_f.to_markdown(index=False) 
            context_text += "## Réconciliation Ligne-à-Ligne - Fournisseurs (Seules les lignes avec écart sont affichées)\n"
            context_text += "```markdown\n" + df_f_str + "\n```\n\n"
        except NameError:
             context_text += "## Réconciliation Ligne-à-Ligne - Fournisseurs\n(Données détaillées non incluses : la librairie 'tabulate' est manquante)\n\n"
        
    if df_reco_c is not None and not df_reco_c.empty:
        try:
            df_c_str = df_reco_c.to_markdown(index=False)
            context_text += "## Réconciliation Ligne-à-Ligne - Clients (Seules les lignes avec écart sont affichées)\n"
            context_text += "```markdown\n" + df_c_str + "\n```\n\n"
        except NameError:
             context_text += "## Réconciliation Ligne-à-Ligne - Clients\n(Données détaillées non incluses : la librairie 'tabulate' est manquante)\n\n"
        
    return context_text

def generate_analysis_id(sage_f, erp_f, sage_c, erp_c, date_deb, date_fin):
    """Génère un identifiant unique (hash) pour l'analyse basée sur les totaux et les dates."""
    data_str = (
        json.dumps(sage_f) + json.dumps(erp_f) + 
        json.dumps(sage_c) + json.dumps(erp_c) + 
        date_deb.isoformat() + date_fin.isoformat()
    )
    return hashlib.sha256(data_str.encode('utf-8')).hexdigest()

def get_total_summary(sage_f_nums, erp_f_nums, sage_c_nums, erp_c_nums):
    """Génère un résumé textuel pour l'historique."""
    summary = []
    
    # Calcul des écarts nets sur le solde final
    f_net_ecart = abs((sage_f_nums[4] - sage_f_nums[5]) - (erp_f_nums[4] - erp_f_nums[5]))
    c_net_ecart = abs((sage_c_nums[4] - sage_c_nums[5]) - (erp_c_nums[4] - erp_c_nums[5]))

    if any(v != 0 for v in sage_f_nums + erp_f_nums) and f_net_ecart >= 0.01:
        summary.append(f"F: Ecart Solde Final Net: {fmt_number(f_net_ecart)} FCFA")
    
    if any(v != 0 for v in sage_c_nums + erp_c_nums) and c_net_ecart >= 0.01:
        summary.append(f"C: Ecart Solde Final Net: {fmt_number(c_net_ecart)} FCFA")
        
    if not summary and (any(v != 0 for v in sage_f_nums + erp_f_nums) or any(v != 0 for v in sage_c_nums + erp_c_nums)):
        return "Réconciliation parfaite (Ecarts nets < 0,01 FCFA)."
        
    return " | ".join(summary) if summary else "Aucune donnée de totaux significative."

# --- FONCTION MODIFIÉE : AJOUT DE LA PERSISTENCE ---
def save_current_analysis(
    sage_f_nums, erp_f_nums, ecarts_f, df_reconciliation_f, 
    sage_c_nums, erp_c_nums, ecarts_c, df_reconciliation_c, 
    date_debut, date_fin, extraction_date
):
    """Enregistre l'analyse complète (Totaux + Tableaux + Recherche des Écarts + Réconciliation ligne par ligne)."""
    
    if not (any(v != 0 for v in sage_f_nums + erp_f_nums) or any(v != 0 for v in sage_c_nums + erp_c_nums)):
        st.warning("Aucune donnée chiffrée de totaux significative à enregistrer.")
        return False
        
    unique_id = generate_analysis_id(sage_f_nums, erp_f_nums, sage_c_nums, erp_c_nums, date_debut, date_fin)

    



# ---------------------------
# Période & uploads
# ---------------------------
st.markdown("---")
st.markdown("#### 🗓 Import des fichiers et période d'analyse")
# ... (logique d'upload des fichiers)
col1, col2 = st.columns(2)
with col1:
    erp_fournisseurs = st.file_uploader("ERP Fournisseurs", type=["xlsx", "csv"], key="erp_fournisseurs")
with col2:
    sage_fournisseurs = st.file_uploader("SAGE Fournisseurs", type=["xlsx", "csv"], key="sage_fournisseurs")
col3, col4 = st.columns(2)
with col3:
    erp_clients = st.file_uploader("ERP Clients", type=["xlsx", "csv"], key="erp_clients")
with col4:
    sage_clients = st.file_uploader("SAGE Clients", type=["xlsx", "csv"], key="sage_clients")


start_date_extracted, end_date_extracted, extraction_date_extracted = None, None, None
if sage_fournisseurs:
    start_date_extracted, end_date_extracted, extraction_date_extracted = extract_all_dates_from_file(sage_fournisseurs)
elif erp_fournisseurs:
    start_date_extracted, end_date_extracted, extraction_date_extracted = extract_all_dates_from_file(erp_fournisseurs)
elif sage_clients:
    start_date_extracted, end_date_extracted, extraction_date_extracted = extract_all_dates_from_file(sage_clients)
elif erp_clients:
    start_date_extracted, end_date_extracted, extraction_date_extracted = extract_all_dates_from_file(erp_clients)

default_end = end_date_extracted or date.today()
default_start = start_date_extracted or (default_end - timedelta(days=30))
default_extraction = extraction_date_extracted or date.today()

col1, col2 = st.columns(2)
with col1:
    date_debut = st.date_input("Date de début", value=default_start, key='date_debut_global')
with col2:
    date_fin = st.date_input("Date de fin", value=default_end, key='date_fin_global')

if date_debut > date_fin:
    st.error("❌ La date de début ne peut pas être postérieure à la date de fin.")
    st.stop()

extraction_date = st.date_input("Date des données", value=default_extraction, key="date_tiree")

# ---------------------------
# État de l'application (Globaux pour l'IA et l'Historique)
# ---------------------------
if 'sage_nums_f' not in st.session_state: st.session_state['sage_nums_f'] = [0.0] * 6
if 'erp_nums_f' not in st.session_state: st.session_state['erp_nums_f'] = [0.0] * 6
if 'ecarts_f' not in st.session_state: st.session_state['ecarts_f'] = [0.0] * 6
if 'sage_nums_c' not in st.session_state: st.session_state['sage_nums_c'] = [0.0] * 6
if 'erp_nums_c' not in st.session_state: st.session_state['erp_nums_c'] = [0.0] * 6
if 'ecarts_c' not in st.session_state: st.session_state['ecarts_c'] = [0.0] * 6
if 'df_reconciliation_f' not in st.session_state: st.session_state['df_reconciliation_f'] = pd.DataFrame()
if 'df_reconciliation_c' not in st.session_state: st.session_state['df_reconciliation_c'] = pd.DataFrame()
if "messages" not in st.session_state: st.session_state["messages"] = []
if "ia_context" not in st.session_state: st.session_state["ia_context"] = ""
# NOUVELLE INITIALISATION : Chargement de l'historique depuis le fichier
if "selected_tab" not in st.session_state: st.session_state["selected_tab"] = "Totaux"




# ---------------------------
# Onglets Streamlit
# ---------------------------
tabs_list = ["Totaux", "Tableaux", "🔍 Recherche des écarts", "🎯 Réconciliation Ligne-à-Ligne", "🤖 Interprétation IA"]

 
tab1, tab2, tab_écarts, tab_reconciliation, tab_ia = st.tabs(tabs_list)

# ... (contenu des onglets)
cleaned_files = {}

# --- Contenu des onglets (omissions pour la concision - voir code précédent) ---

with tab1:
    st.subheader("📊 Résumé des Totaux")
    def display_simple_totals(entity_name, sage_nums, erp_nums):
        col1, col2 = st.columns(2)
        erp_solde_init_net = abs(erp_nums[0] - erp_nums[1])
        erp_mvt_d = erp_nums[2]
        erp_mvt_c = erp_nums[3]
        erp_solde_final_net = abs(erp_nums[4] - erp_nums[5])
        sage_solde_init_net = abs(sage_nums[0] - sage_nums[1])
        sage_mvt_d = sage_nums[2]
        sage_mvt_c = sage_nums[3]
        sage_solde_final_net = abs(sage_nums[4] - sage_nums[5])
        with col1:
            st.markdown(f"#### ERP {entity_name}")
            st.write(f"Solde d'ouverture : {fmt_number(erp_solde_init_net)}")
            st.write(f"Mvt débit : {fmt_number(erp_mvt_d)}")
            st.write(f"Mvt crédit : {fmt_number(erp_mvt_c)}")
            st.write(f"Solde final : {fmt_number(erp_solde_final_net)}")
            erp_pie_values = [abs(erp_solde_init_net), erp_mvt_d, erp_mvt_c, abs(erp_solde_final_net)]
            erp_pie_labels = ["Solde Ouv. NET", "Mvt Débit", "Mvt Crédit", "Solde Final NET"]
            if any(val > 0 for val in erp_pie_values):
                create_pie_chart(f"Répartition des Totaux - ERP {entity_name}", erp_pie_values, erp_pie_labels)
            else:
                st.info(f"Aucune donnée chiffrée à afficher pour ERP {entity_name}.")
        with col2:
            st.markdown(f"#### SAGE {entity_name}")
            st.write(f"Solde d'ouverture : {fmt_number(sage_solde_init_net)}")
            st.write(f"Mvt débit : {fmt_number(sage_mvt_d)}")
            st.write(f"Mvt crédit : {fmt_number(sage_mvt_c)}")
            st.write(f"Solde final : {fmt_number(sage_solde_final_net)}")
            sage_pie_values = [abs(sage_solde_init_net), sage_mvt_d, sage_mvt_c, abs(sage_solde_final_net)]
            sage_pie_labels = ["Solde Ouv.", "Mvt Débit", "Mvt Crédit", "Solde Final"]
            if any(val > 0 for val in sage_pie_values):
                create_pie_chart(f"Répartition des Totaux - SAGE {entity_name}", sage_pie_values, sage_pie_labels)
            else:
                st.info(f"Aucune donnée chiffrée à afficher pour SAGE {entity_name}.")

    wrote_any = False
    if erp_fournisseurs or sage_fournisseurs:
        sage_nums_f, _, erp_nums_f, _ = compute_totals("Fournisseurs", erp_fournisseurs, sage_fournisseurs)
        st.session_state['sage_nums_f'] = sage_nums_f
        st.session_state['erp_nums_f'] = erp_nums_f
        display_simple_totals("Fournisseurs", sage_nums_f, erp_nums_f)
        wrote_any = True
    if erp_clients or sage_clients:
        st.markdown("---")
        sage_nums_c, _, erp_nums_c, _ = compute_totals("Clients", erp_clients, sage_clients)
        st.session_state['sage_nums_c'] = sage_nums_c
        st.session_state['erp_nums_c'] = erp_nums_c
        display_simple_totals("Clients", sage_nums_c, erp_nums_c)
        wrote_any = True
    if not wrote_any:
        st.info("Veuillez importer des fichiers pour voir le résumé.")


with tab2:
    st.subheader("📋 Tableaux détaillés et Histogrammes")
    wrote_any = False
    chart_labels_nets = ["Solde d'ouverture", "Mouvements", "Solde Final"]
    if erp_fournisseurs or sage_fournisseurs:
        st.markdown("---")
        st.markdown("#### COMPTABILISATION DES OPERATIONS SUR LES FOURNISSEURS")
        sage_nums_f, erp_nums_f, ecarts_f, gaps_f = render_html_table("Fournisseurs", erp_fournisseurs, sage_fournisseurs, extraction_date, date_debut, date_fin)
        st.session_state['ecarts_f'] = ecarts_f
        st.session_state['sage_nums_f'] = sage_nums_f
        st.session_state['erp_nums_f'] = erp_nums_f
        if any(v != 0 for v in sage_nums_f) or any(v != 0 for v in erp_nums_f):
            st.markdown("---")
            create_bar_chart("Écarts Fournisseurs (SAGE vs ERP)", chart_labels_nets, sage_nums_f, erp_nums_f, ecarts_f, gaps_f)
        wrote_any = True
    if erp_clients or sage_clients:
        st.markdown("---")
        st.markdown("#### COMPTABILISATION DES OPERATIONS SUR LES CLIENTS")
        sage_nums_c, erp_nums_c, ecarts_c, gaps_c = render_html_table("Clients", erp_clients, sage_clients, extraction_date, date_debut, date_fin)
        st.session_state['ecarts_c'] = ecarts_c
        st.session_state['sage_nums_c'] = sage_nums_c
        st.session_state['erp_nums_c'] = erp_nums_c
        if any(v != 0 for v in sage_nums_c) or any(v != 0 for v in erp_nums_c):
            st.markdown("---")
            create_bar_chart("Écarts Clients (SAGE vs ERP)", chart_labels_nets, sage_nums_c, erp_nums_c, ecarts_c, gaps_c)
        wrote_any = True
    if wrote_any:
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Totaux"
        next_row = 1
        if erp_fournisseurs or sage_fournisseurs:
            next_row = write_styled_table(ws, next_row, "Fournisseurs", erp_fournisseurs, sage_fournisseurs, extraction_date, date_debut, date_fin, section_title="COMPTABILISATION DES OPERATIONS SUR LES FOURNISSEURS")
        if erp_clients or sage_clients:
            next_row += 1
            next_row = write_styled_table(ws, next_row, "Clients", erp_clients, sage_clients, extraction_date, date_debut, date_fin, section_title="COMPTABILISATION DES OPERATIONS SUR LES CLIENTS")
        wb.save(output)
        st.download_button(
            label="📥 Télécharger Excel",
            data=output.getvalue(),
            file_name=f"Totaux_SAGE_ERP_{date.today().strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Veuillez importer des fichiers pour voir les tableaux et graphiques.")



    def clean_excel_file(uploaded_file, source="sage"):
        if uploaded_file is None:
            return None 
        try:
            ext = uploaded_file.name.split('.')[-1].lower()
            if ext == "xlsx":
                df_raw = pd.read_excel(uploaded_file, engine="openpyxl", header=None)
            elif ext == "csv":
                uploaded_file.seek(0)
                df_raw = pd.read_csv(uploaded_file, encoding="latin-1", on_bad_lines="skip", low_memory=False, header=None)
            else:
                st.error("Format de fichier non supporté")
                return None

            # 🔹 Détection automatique de la ligne d'en-têtes
            header_row = None
            for i, row in df_raw.iterrows():
                row_str = " ".join(str(x) for x in row.tolist()).lower()
                if "débit" in row_str and "crédit" in row_str:
                    header_row = i
                    break
            if header_row is None:
                st.warning("Impossible de trouver la ligne d'en-têtes (Débit/Crédit)")
                return None

            # 🔹 Relecture propre
            uploaded_file.seek(0)
            if ext == "xlsx":
                df = pd.read_excel(uploaded_file, engine="openpyxl", header=header_row)
            else:
                df = pd.read_csv(uploaded_file, encoding="latin-1", on_bad_lines="skip", low_memory=False, header=header_row)

            # 🔹 Nettoyage global
            df = df[~df.apply(lambda row: row.astype(str).str.lower().str.contains("total").any(), axis=1)]
            df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)
            df = df.loc[:, df.columns.notna()].fillna(0)

            # 🔹 Conversion des valeurs
            for col in df.columns:
                if "débit" in str(col).lower() or "crédit" in str(col).lower():
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            # 🔹 Renommage des colonnes selon la source
            if source == "sage":
                new_cols = ["Numéro de compte", "Intitulé des comptes", "Débit 1", "Crédit 1", "Débit 2", "Crédit 2", "Débit 3", "Crédit 3"]
            else:
                new_cols = ["Compte ERP", "Nom ERP", "Débit 1", "Crédit 1", "Débit 2", "Crédit 2", "Débit 3", "Crédit 3"]
            df.columns = new_cols[:len(df.columns)]

            return df
        except Exception as e:
            st.error(f"Erreur lors du nettoyage du fichier : {e}")
            return None

    # --- Chargement et affichage des fichiers ---
    files_dict = {
        "SAGE Fournisseurs": (sage_fournisseurs, "sage"),
        "ERP Fournisseurs": (erp_fournisseurs, "erp"),
        "SAGE Clients": (sage_clients, "sage"),
        "ERP Clients": (erp_clients, "erp")
    }

    for key, (f, source) in files_dict.items():
        if f:
            df_clean = clean_excel_file(f, source)
            if df_clean is not None:
                st.markdown(f"### 📂 {key} - Aperçu après nettoyage")
                st.dataframe(df_clean.head(15))
                cleaned_files[key] = df_clean



with tab_écarts:
    st.subheader("🔍 Recherche des Écarts")
    
    def run_gap_analysis(entity_type, sage_key, erp_key):
        # Logique d'analyse des écarts
        st.markdown("---")
        st.markdown(f"## {entity_type} : Correspondances SAGE vs ERP")
        
        sage_f = cleaned_files.get(sage_key)
        erp_f = cleaned_files.get(erp_key)
        
        if sage_f is None or erp_f is None:
            st.info(f"Veuillez importer et nettoyer les fichiers {sage_key} et {erp_key}.")
            return
            
        sage_col = sage_f.columns[0]
        erp_col = erp_f.columns[0]
        correspondances = []
        non_correspondances_sage = []
        erp_rows_used = set()
        
        def split_blocks(s): 
            return re.findall(r'\d+|[A-Za-z]+', str(s).strip())
        
        def is_match(sage_str, erp_str):
            sage_blocks = split_blocks(sage_str)
            erp_blocks = split_blocks(erp_str)
            if not sage_blocks or not erp_blocks: 
                return False
            idx = 0
            for block in sage_blocks:
                while idx < len(erp_blocks) and block != erp_blocks[idx]: 
                    idx += 1
                if idx == len(erp_blocks): 
                    return False
                idx += 1
            return True

        for _, sage_row in sage_f.iterrows():
            sage_str = str(sage_row[sage_col]).strip()
            match_found = False
            for erp_idx, erp_row in erp_f.iterrows():
                if erp_idx in erp_rows_used: 
                    continue
                erp_str = str(erp_row[erp_col]).strip()
                if is_match(sage_str, erp_str) or is_match(erp_str, sage_str):
                    merged_row = pd.concat([sage_row, erp_row], axis=0)
                    correspondances.append(merged_row)
                    erp_rows_used.add(erp_idx)
                    match_found = True
                    break
            if not match_found: 
                non_correspondances_sage.append(sage_row)
            
        if correspondances:
            df_correspondances = pd.concat(correspondances, axis=1).T.reset_index(drop=True)
            cols = pd.Series(df_correspondances.columns)
            for col in cols:
                if "Débit" in str(col) or "Crédit" in str(col):
                    indices = cols[cols == col].index.values.tolist()
                    if len(indices) >= 2:
                        cols.iloc[indices[0]] = f"{col} (SAGE)"
                        cols.iloc[indices[1]] = f"{col} (ERP)"
            df_correspondances.columns = cols
            st.markdown(f"### 📌 {entity_type} SAGE vs ERP correspondants")
            st.dataframe(df_correspondances)
        else:
            st.info(f"Aucune correspondance trouvée entre SAGE et ERP pour les {entity_type}.")
            
        if non_correspondances_sage:
            st.markdown(f"### ❌ {entity_type} SAGE sans correspondance ERP")
            df_non_correspondances = pd.DataFrame(non_correspondances_sage)
            st.dataframe(df_non_correspondances)
            
        non_correspondances_erp = erp_f.loc[~erp_f.index.isin(erp_rows_used)]
        if not non_correspondances_erp.empty:
            st.markdown(f"### ❌ {entity_type} ERP sans correspondance SAGE")
            st.dataframe(non_correspondances_erp)

        # 🔹 Sauvegarde des résultats d’écarts dans la session (pour historique)
        if entity_type == "Fournisseurs":
            st.session_state["df_gap_f"] = df_correspondances if 'df_correspondances' in locals() else pd.DataFrame()
        elif entity_type == "Clients":
            st.session_state["df_gap_c"] = df_correspondances if 'df_correspondances' in locals() else pd.DataFrame()

    # --- Appels de la fonction pour les deux types ---
    run_gap_analysis("Fournisseurs", "SAGE Fournisseurs", "ERP Fournisseurs")
    run_gap_analysis("Clients", "SAGE Clients", "ERP Clients")




from openpyxl import load_workbook
from openpyxl.styles import Border, Side


with tab_reconciliation:
    st.subheader("🎯 Réconciliation détaillée - Écarts individuels (SAGE - ERP)")

    def run_reconciliation(entity_type, sage_key, erp_key, session_state_key):
        st.markdown("---")
        st.markdown(f"## {entity_type} : Analyse des Écarts Ligne-à-Ligne")
        
        sage_f = cleaned_files.get(sage_key)
        erp_f = cleaned_files.get(erp_key)
        
        if sage_f is None or erp_f is None:
            st.info(f"Veuillez d'abord importer et nettoyer les fichiers SAGE et ERP {entity_type} dans l'onglet 'Nettoyage Excel'.")
            st.session_state[session_state_key] = pd.DataFrame()
            return
        
        sage_col = sage_f.columns[0]
        erp_col = erp_f.columns[0]
        correspondances = []
        erp_rows_used = set()
        
        def split_blocks(s): 
            return re.findall(r'\d+|[A-Za-z]+', str(s).strip())
        
        def is_match(sage_str, erp_str):
            sage_blocks = split_blocks(sage_str)
            erp_blocks = split_blocks(erp_str)
            if not sage_blocks or not erp_blocks:
                return False
            idx = 0
            for block in sage_blocks:
                while idx < len(erp_blocks) and block != erp_blocks[idx]:
                    idx += 1
                if idx == len(erp_blocks):
                    return False
                idx += 1
            return True

        # --- Correspondances ---
        for _, sage_row in sage_f.iterrows():
            sage_str = str(sage_row[sage_col]).strip()
            for erp_idx, erp_row in erp_f.iterrows():
                if erp_idx in erp_rows_used:
                    continue
                erp_str = str(erp_row[erp_col]).strip()
                if is_match(sage_str, erp_str) or is_match(erp_str, sage_str):
                    merged_row = pd.concat([sage_row, erp_row], axis=0)
                    correspondances.append(merged_row)
                    erp_rows_used.add(erp_idx)
                    break
        
        if correspondances:
            df_correspondances = pd.concat(correspondances, axis=1).T.reset_index(drop=True)
            cols = pd.Series(df_correspondances.columns)
            
            for dup in cols[cols.duplicated()].unique():
                if "Débit" in str(dup) or "Crédit" in str(dup):
                    indices = cols[cols == dup].index.values.tolist()
                    if len(indices) >= 2:
                        cols.iloc[indices[0]] = f"{dup}.SAGE"
                        cols.iloc[indices[1]] = f"{dup}.ERP"
            df_correspondances.columns = cols
            
            df_reconciliation = calculate_detailed_gaps(df_correspondances, entity_type)
            st.session_state[session_state_key] = df_reconciliation

            display_reconciliation_table(df_reconciliation, f"Détail des Écarts {entity_type}")

            # --- 🔽 Bouton Télécharger Excel avec bordures ---
            if not df_reconciliation.empty:
                output = BytesIO()

                # Étape 1 : écrire le DataFrame dans Excel
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_reconciliation.to_excel(writer, index=False, sheet_name=entity_type[:31])

                output.seek(0)

                # Étape 2 : ouvrir avec openpyxl pour appliquer les bordures
                wb = load_workbook(output)
                ws = wb.active

                # Définir le style de bordure fine
                thin_border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )

                # Appliquer la bordure à toutes les cellules remplies
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border

                # Étape 3 : sauvegarder de nouveau dans un buffer
                bordered_output = BytesIO()
                wb.save(bordered_output)
                bordered_output.seek(0)

                # Étape 4 : bouton de téléchargement
                st.download_button(
                    label=f"📥 Télécharger {entity_type}",
                    data=bordered_output,
                    file_name=f"Reconciliation_{entity_type}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

        else:
            st.info(f"Aucune correspondance {entity_type} trouvée pour l'analyse des écarts.")
            st.session_state[session_state_key] = pd.DataFrame()

    # --- Exécution pour chaque entité ---
    run_reconciliation("Fournisseurs", "SAGE Fournisseurs", "ERP Fournisseurs", 'df_reconciliation_f')
    run_reconciliation("Clients", "SAGE Clients", "ERP Clients", 'df_reconciliation_c')


# ---------------------------
# Onglet : Interprétation IA 
# ---------------------------
with tab_ia:
    st.subheader("🤖 Assistant IA - Interprétation de la Réconciliation")
    # Logique IA... (inchangée)
    if not MODEL_NAME:
        st.info("Veuillez configurer la clé API Gemini dans le fichier secrets.toml pour activer l'assistant.")
        # st.stop() # Ne pas stopper ici car cela casserait le reste de l'application
    else:
        client = st.session_state.get(CLIENT_KEY)
        if client is None:
            st.error("Erreur: Le client Gemini n'a pas pu être initialisé. Vérifiez votre clé API et l'état de la session.")
        else:
            current_context = generate_full_context(
                st.session_state['sage_nums_f'], st.session_state['erp_nums_f'], st.session_state['ecarts_f'],
                st.session_state['sage_nums_c'], st.session_state['erp_nums_c'], st.session_state['ecarts_c'],
                st.session_state['df_reconciliation_f'], st.session_state['df_reconciliation_c']
            )
            context_changed = st.session_state.get("ia_context", "") != current_context
            if context_changed:
                st.session_state.messages = []
                st.session_state["ia_context"] = current_context
            system_instruction = (
                "Vous êtes un expert-comptable assistant spécialisé dans la réconciliation des balances auxiliaires (Fournisseurs/Clients) "
                "entre le système SAGE et un ERP. Votre objectif est d'analyser les données de réconciliation fournies "
                "et de répondre aux questions de l'utilisateur de manière claire, concise et professionnelle. "
                "Les valeurs sont exprimées en FCFA (Francs CFA). "
                "Lorsque vous citez des données tabulaires, utilisez le format Markdown pour les rendre lisibles. "
                "Le contexte d'analyse actuel, y compris les totaux et les écarts ligne-à-ligne (si fournis), est le suivant:\n\n" 
                + current_context
            )
            if not st.session_state.messages:
                st.info("Assistant initialisé. Le contexte d'analyse (totaux, écarts) a été chargé. Posez votre première question (ex: 'Quel est l'écart sur le solde final des fournisseurs ?').")
                
            for message in st.session_state.messages:
                with st.chat_message(message["role"]):
                    st.markdown(message["content"])
            
            if prompt := st.chat_input("Posez une question sur les écarts, les totaux, ou les lignes de réconciliation..."):
                st.session_state.messages.append({"role": "user", "content": prompt})
                with st.chat_message("user"):
                    st.markdown(prompt)
                history_contents = []
                for msg in st.session_state.messages:
                     role = 'user' if msg['role'] == 'user' else 'model'
                     history_contents.append(types.Content(role=role, parts=[types.Part(text=msg['content'])]))
                config = types.GenerateContentConfig(
                    system_instruction=system_instruction
                )
                with st.chat_message("assistant"):
                    with st.spinner("🤖 Assistant en cours de réflexion..."):
                        try:
                            response = client.models.generate_content(
                                 model=MODEL_NAME,
                                 contents=history_contents,
                                 config=config
                            )
                            st.markdown(response.text)
                            st.session_state.messages.append({"role": "assistant", "content": response.text})
                        except APIError as e:
                            st.error(f"Erreur de l'API Gemini: {e}")
                            st.session_state.messages.pop() 
                        except Exception as e:
                            st.error(f"Une erreur inattendue est survenue: {e}")
                            st.session_state.messages.pop()


# ---------------------------
# Onglet : Historique (MODIFIÉ)
# ---------------------------
